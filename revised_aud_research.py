# -*- coding: utf-8 -*-
"""
Created on Wed Oct 27 14:24:26 2021

@author: gsmittkamp
"""

import os, re, sys, shutil, datetime, warnings, time
import win32com.client as win32
import pandas as pd
import numpy as np
from tabulate import tabulate
from openpyxl import load_workbook
warnings.filterwarnings("ignore")
import new_selections
from gs_utility import clear_excel_cache
#import qc_aud_prod
from tqdm import tqdm


def read_audiences(shared_wb,option):
    grant_projects=shared_wb
    
    if option == 'mb':
        grant_projects=grant_projects[grant_projects.Status=='Model Built']
    elif option == 'sr':
        grant_projects=grant_projects[grant_projects.Status=='Selection Ready']
    elif option == 'sb':
        grant_projects=grant_projects[grant_projects.Status=='Selection Begun']
    elif option == 'sa':
        grant_projects=grant_projects[grant_projects.Status=='Selection Assigned']
    elif option == 'pr':
        grant_projects=grant_projects[grant_projects.Status=='Post-Selection S&O Review']
    elif option == 'pa':
        grant_projects=grant_projects[grant_projects.Status=='Post-Selection S&O Approval']
    elif option == 'rs':
        grant_projects=grant_projects[grant_projects.Status=='Ready to Ship']

    return grant_projects


def create_audience_info(grant_projects):
    combined_audience_workbook=pd.read_excel(r'C:\Users\gsmittkamp\taxonomy_and_wave_workbook.xlsx', sheet_name='combined')
    merged=pd.merge(grant_projects,combined_audience_workbook,how='left',left_on='Audience Display ID',right_on='Audience_Display_ID')
    merged['Audience Name']=np.where(pd.isna(merged['Platform Audience Name']),merged['Wave_Audience_Name'],merged['Platform Audience Name'])
    merged['Market_Niche_Description']=np.where(pd.isna(merged['Taxonomic Description']),merged['Wave_Market_Niche_Description'],merged['Taxonomic Description'])
    merged=merged[['Audience Display ID', 'Market Niche ID','Audience Name','Status','Selection Request ID','Market_Niche_Description','Exemplars','Score ID', 'Market Size','Model Request ID', 'S&O Notes']]
    merged.columns = merged.columns.str.replace(' ', '_')
    merged.sort_values(by=['Audience_Display_ID'],ascending=True, inplace=True)
    merged.rename(columns={'S&O_Notes':'Current_S&O_Notes'}, inplace=True)
    return merged


def select_omits(merged):
    income=pd.read_excel(r'C:\Users\gsmittkamp\audience_size.xlsx', sheet_name='income_omits')
    gender=pd.read_excel(r'C:\Users\gsmittkamp\audience_size.xlsx', sheet_name='gender_omits')
    merged['potential_so_notes']=''
    merged['income_so_notes']=''
    merged['gender_so_notes']=''
    for keyword,note in zip(income.keyword,income.omit):
        merged.income_so_notes=np.where(merged.Market_Niche_Description.str.lower().str.contains(keyword),note,merged.income_so_notes)
        merged.income_so_notes=np.where((merged['Audience_Name'].str.lower().str.contains(keyword))&(merged.income_so_notes==''),note,merged.income_so_notes)
    for keyword,note in zip(gender.keyword,gender.omit):
        merged.gender_so_notes=np.where(merged.Market_Niche_Description.str.lower().str.contains(keyword),note,merged.gender_so_notes)
        merged.gender_so_notes=np.where((merged['Audience_Name'].str.lower().str.contains(keyword))&(merged.gender_so_notes==''),note,merged.gender_so_notes)
    merged.potential_so_notes=np.where((merged.income_so_notes!='')&(merged.gender_so_notes==''),merged.income_so_notes,merged.potential_so_notes)
    merged.potential_so_notes=np.where((merged.income_so_notes=='')&(merged.gender_so_notes!=''),merged.gender_so_notes,merged.potential_so_notes)
    merged.potential_so_notes=np.where((merged.income_so_notes!='')&(merged.gender_so_notes!=''),merged.income_so_notes+", "+merged.gender_so_notes,merged.potential_so_notes)
    merged.potential_so_notes=np.where(merged.potential_so_notes=='','none',merged.potential_so_notes)
    merged=merged[['Audience_Display_ID', 'Market_Niche_ID', 'Audience_Name', 'Status', 'Website_Display_Name', 'Market_Niche_Description', 'Exemplars', 'Score_ID', 'Market_Size', 'Model_Request_ID', 'Selection_Request_ID', 'S&O_Notes', 'potential_so_notes']]
    merged.rename(columns={'S&O_Notes':'Current_S&O_Notes'}, inplace=True)
    return merged


def check_market_size(merged):
    ## put these in with omit keylist at some point
    #market_size=pd.read_excel(r'C:\Users\gsmittkamp\audience_size.xlsx', sheet_name='size')
    audience_size=merged[['Audience_Display_ID','Audience_Name','Market_Niche_Description','Market_Size','Model_Request_ID', 'Selection_Request_ID','Current_S&O_Notes']]
    audience_size['Correct_size']=''
    #audience_size['Correct_size']=np.where(audience_size['Audience_Display_ID'].str.contains('UAA'),audience_size.UAA_Size,audience_size.Correct_size)
    #audience_size['Correct_size']=np.where(audience_size['Audience_Display_ID'].str.contains('USP'),audience_size.USP_Size,audience_size.Correct_size)
    #if audiences are in pre-selection there will not be a selection request_ID yet, so handle those cases by setting to integer of 0 and then creating URL for it if actually has a real selection request_id
    audience_size['Selection_Request_ID']=audience_size.Selection_Request_ID.fillna(value=0).astype(int)
    audience_size['URL']=np.where(audience_size.Selection_Request_ID!=0,"https://intranet.wiland.com/CountRequest/DigitalAudiences/ShipOrder.php?srid="+(audience_size.Selection_Request_ID.astype(int)).astype(str),'')
    audience_size['Shipped?']=''
    audience_size=audience_size[["Audience_Display_ID", "Selection_Request_ID", "Audience_Name", "Shipped?", "URL", "Correct_size", "Current_S&O_Notes"]]
    
    #check for any duplicate selection requests to avoid shipping same request more than once
    def check_duplicate_reqs(audience_size):
        agg_df=pd.DataFrame(audience_size.groupby('Selection_Request_ID').apply(len)).reset_index()
        dupes=agg_df[agg_df[0]>1]
        if len(dupes)>0:
            for request in dupes.Selection_Request_ID:
                if int(request) != 0:
                    input('''Duplicate selection requests exist. Please fix duplicate requests and re-start program \n\n'''+str(request))
            
    check_duplicate_reqs(audience_size)
    
    return audience_size


def create_header_data(merged):
    header_data=merged[['Audience_Display_ID','Audience_Name','Market_Niche_ID','Market_Niche_Description','Current_S&O_Notes']]
    header_data['Current_S&O_Notes'][header_data['Current_S&O_Notes']=='none']=''
    header_data['SA Team Member']='Grant Smittkamp'
    header_data=header_data[['Audience_Display_ID','Audience_Name','Market_Niche_ID','Market_Niche_Description','SA Team Member','Current_S&O_Notes']]
    header_data.sort_values(by=['Audience_Display_ID'],ascending=True, inplace=True)
    header_data.set_index('Audience_Display_ID', inplace=True)
    header_data=header_data.T
    return header_data


def read_Wiland_App_attributess(merged):
    
    #each time selections are assigned, download the audience attributes from app and this function will find which filename is most recent from downloads, and use it to calculate audience attributes to use
    def find_new_attribute_file():
        onlyfiles = [f for f in os.listdir(r"C:\Users\gsmittkamp") if os.path.isfile(os.path.join(r"C:\Users\gsmittkamp", f))]
        attribute_files = pd.DataFrame({"File_Name": [],"Date": []})
        for file_name in onlyfiles:
            if file_name != '' and "Audience_Attributes" in file_name:
                date_segment=file_name.split('_')[2]
                date=date_segment.split('.')[0].replace('-','/')
                attribute_files=attribute_files.append({"File_Name": file_name,"Date": date},ignore_index=True)
        attribute_files.Date=pd.to_datetime(attribute_files.Date,errors='coerce')
        newest = attribute_files['Date'].max()
        #test to make sure the newest attribute file possible is being used
        confirm_today=datetime.datetime.today().strftime('%B %d, %Y')
        confirm_newest = newest.strftime('%B %d, %Y')
        return newest, confirm_today, confirm_newest, attribute_files
    
    newest, confirm_today, confirm_newest, attribute_files=find_new_attribute_file()
    
    if confirm_today != confirm_newest:
        pause_option = input('''Not using newest attribute file, continue anyways? (y/n) \n\nOption: ''')
        newest, confirm_today, confirm_newest, attribute_files=find_new_attribute_file()
        
        if confirm_today != confirm_newest:
            pause_option = input('''Still not using newest attribute file, download then confirm? (y/n) \n\nOption: ''')
            newest, confirm_today, confirm_newest, attribute_files=find_new_attribute_file()
      
    print("using attribute file for "+str(newest))
    newest_file=attribute_files.File_Name[attribute_files.Date==newest]
    newest_file_name=newest_file.loc[max(newest_file.index)]
    current_attributes=pd.read_excel(r"C:\Users\gsmittkamp\\"+newest_file_name)
    
    #some audiences may still be in selection, so we do not want to proceed until selection has completeed
    merged['in_selection']=np.where(~merged['Audience_Display_ID'].isin(current_attributes['Audience Display ID']),True,False)
    if sum(merged['in_selection'])>0:
        #create a printable string to inform user which audiences cannot currently proceed with post-selection
        in_selection=merged[merged['in_selection']]
        selection_running_str=''
        counter=0
        for aud in in_selection['Audience_Display_ID']:
            counter+=1
            print(aud+" still in selection")
            if counter < len(in_selection):
                selection_running_str+=aud+", "
            else:
                selection_running_str+=aud
            
        forced_pause=input('''\nAttributes not yet ready for %s of the %s audiences\n\nProceed with other audiences?:'''%(str(len(in_selection)), str(len(merged))))
        if forced_pause.lower()=='y':
            merged=merged[merged['in_selection']==False]
            print(selection_running_str)
        elif forced_pause.lower()=='n':
            quit()
            
    current_attributes=current_attributes[current_attributes['Audience Display ID'].isin(merged['Audience_Display_ID'])]
    current_attributes=pd.merge(current_attributes,merged[['Current_S&O_Notes','Audience_Display_ID']],how='left',left_on='Audience Display ID',right_on='Audience_Display_ID')
    current_attributes['Current_S&O_Notes']=np.where(current_attributes['Current_S&O_Notes']=='none','',current_attributes['Current_S&O_Notes'])
    current_attributes=current_attributes[['Audience Display ID','Audience Name','Percent Female','Percent Male','Married Percentage','Percent w/ Children at Home','Median Age','Median HouseHold Income','Current_S&O_Notes']]
    current_attributes=current_attributes.round({'Married Percentage': 0, 'Percent w/ Children at Home': 0})
    current_attributes['Percent Female']=current_attributes['Percent Female'].round(0).astype(int).astype(str)+"% Female"
    current_attributes['Percent Male']=current_attributes['Percent Male'].round(0).astype(int).astype(str)+"% Male"
    current_attributes['Married Percentage']=current_attributes['Married Percentage']/100
    current_attributes['Percent w/ Children at Home']=current_attributes['Percent w/ Children at Home']/100
    current_attributes['Median HouseHold Income']=current_attributes['Median HouseHold Income']*100
    current_attributes.sort_values(by=['Audience Display ID'],ascending=True, inplace=True)
    current_attributes.set_index('Audience Display ID', inplace=True)
    normal_attributes=current_attributes.reset_index()
    current_attributes=current_attributes.T
    
    return current_attributes,normal_attributes


def find_current_attributes(merged,normal_attributes,existing_files_df):
    
    def get_all_sheets(current_checklist,aud_id):
        print("getting sheets for "+aud_id)
        sheet_name_df = pd.DataFrame({"Audience Display ID": [],"sheet_name": [],"round": [], "Initials":[], "Approval_Date": [],"Percent Female": [],"Percent Male": [],"Married Percentage": [],"Percent w/ Children at Home": [],"Median Age": [],"Median HouseHold Income": [],"Current_S&O_Notes": []})
        xli = pd.ExcelFile(r'L:\ProductManagement\Standard Audiences\Audience Tracking\Audience Attributes\Omit and Select Checklists\\'+current_checklist)
        sheet_names=list(xli.sheet_names)  # see all sheet names
        sel_round=0
        for sheet in sheet_names:
            def test_sheet_info(sheet):
                sheet_name_string = sheet.lower()
                #sometimes sheet name can be generic, so have to account for that by manually checking and providing the ammended sheet name to the script for use in checking that sheet
                if "pre" not in sheet_name_string and "post" not in sheet_name_string and "post" not in sheet_name_string and "attribute" not in sheet_name_string:
                    format_error=input('''\nsheet with incorrect name in checklist for %s\n\nProceed anyways? (y/n) \n\nOption: :'''%(aud_id))
                    new_sheet_name=input('''What is the corrected sheet name for %s?'''%(aud_id))
                    new_sheet_name=new_sheet_name.strip()
                    sheet=new_sheet_name
                    sheet_name_string = sheet.lower()
                return sheet,sheet_name_string
                
            sheet,sheet_name_string=test_sheet_info(sheet)
                
            if "post selection" in sheet_name_string:
                current_sheet=xli.parse(sheet)
                #if checklist has not been completed for a round of selection skip that sheet
                if pd.isna(current_sheet['Unnamed: 2'].iloc[6]) and pd.isna(current_sheet['Unnamed: 2'].iloc[2]):
                    sheet_name_df=sheet_name_df.append({"Audience Display ID":aud_id,"sheet_name": sheet,"round": sel_round},ignore_index=True)
                else:
                    #test if sheet has values by checking cell for one of the genders
                    #if len(current_sheet.value_1.iloc[22])>1:
                    Initials=str(current_sheet['Unnamed: 2'].iloc[6])
                    if "anthony" in Initials.lower():
                        Initials='AS'
                    if "eric" in Initials.lower():
                        Initials='EH'
                    if "karen" in Initials.lower():
                        Initials='KC'
                    if "emily" in Initials.lower():
                        Initials='EW'
                    if "grant" in Initials.lower():
                        Initials='GS'
                    #if date value is not correct type, ignore and move ont to next via try and except
                    if type(current_sheet['Unnamed: 4'].iloc[7]) == datetime.datetime:
                        Approval_Date=current_sheet['Unnamed: 4'].iloc[7].strftime('%m/%d/%Y')
                    elif current_sheet['Unnamed: 4'].iloc[7] == '':
                        Approval_Date='N/A'
                    #if approval date column has strings, like "none" set to default n/a to avoid errors
                    elif type(current_sheet['Unnamed: 4'].iloc[7]) == str and len(current_sheet['Unnamed: 4'].iloc[7]) >1:
                        Approval_Date='N/A'
                    else:
                        #print("could not parse date for"+aud_id+", with value "+str(current_sheet['Unnamed: 4'].iloc[7]))
                        Approval_Date='N/A'
                    Female=int(current_sheet['Unnamed: 3'].str.rstrip('% Female').iloc[14])/100
                    Male=int(current_sheet['Unnamed: 4'].str.rstrip('% Male').iloc[14])/100
                    Married=current_sheet['Unnamed: 3'].iloc[22]
                    Children=current_sheet['Unnamed: 3'].iloc[30]
                    Med_Age=current_sheet['Unnamed: 3'].iloc[38]
                    Med_HHI=current_sheet['Unnamed: 3'].iloc[46]
                    #old format can somtimes input hhi as hundreds instead of standard units
                    if Med_HHI <= 1500:
                        Med_HHI=Med_HHI*100
                    sel_round+=1
                    def find_omits(current_sheet):
                        current_sheet['Unnamed: 1']=current_sheet['Unnamed: 1'].str.lstrip('Notes:\n ')
                        current_sheet['Unnamed: 1']=current_sheet['Unnamed: 1'].str.lstrip('Notes: ')
                        current_sheet['Unnamed: 1']=current_sheet['Unnamed: 1'].str.lstrip('Already ')
                        current_sheet['Unnamed: 1']=current_sheet['Unnamed: 1'].str.lstrip('pplied')
                        current_sheet['Unnamed: 1']=current_sheet['Unnamed: 1'].str.rstrip(' ')
                        current_sheet['Unnamed: 1']=current_sheet['Unnamed: 1'].str.lstrip(' ')
                        current_sheet['Unnamed: 1']=current_sheet['Unnamed: 1'].str.lstrip('\n')
                        current_sheet['Unnamed: 1']=current_sheet['Unnamed: 1'].str.rstrip('\n')
                        current_sheet['Unnamed: 1']=current_sheet['Unnamed: 1'].str.lstrip('pplied')
                        gender_omit = current_sheet['Unnamed: 1'].iloc[21]
                        married_omit = current_sheet['Unnamed: 1'].iloc[29]
                        children_omit = current_sheet['Unnamed: 1'].iloc[37]
                        age_omit = current_sheet['Unnamed: 1'].iloc[45]
                        hhi_omit = current_sheet['Unnamed: 1'].iloc[53]
                        potential_omits=[gender_omit,married_omit,children_omit,age_omit,hhi_omit]
                        omit_df=pd.DataFrame({"omit": []})
                        omit_count=0
                        omit_notes=''
                        for n in potential_omits:
                            try:
                                n=str(n)
                            except:
                                n='missing'
                            if n != '':
                                omit_count+=1
                                omit_df=omit_df.append({"omit":n},ignore_index=True)
                        #nan's were being passed as strings due to
                        omit_df['omit']=omit_df['omit'].astype(str)
                        omit_df['omit']=omit_df.omit.str.replace('omitted', 'omit')
                        omit_df['omit']=np.where(omit_df.omit.str.contains("nan,"),'',omit_df.omit)
                        if omit_count>1:
                            for n in omit_df.omit:
                                omit_notes+=n+", "
                        elif omit_count==1:
                            for n in omit_df.omit:
                                omit_notes=n
                        omit_notes = omit_notes.strip(" ,")
                        return omit_notes
                    omit_notes=find_omits(current_sheet)
                    if omit_notes == '':
                        omit_notes = 'none'
                    sheet_name_df=sheet_name_df.append({"Audience Display ID":aud_id,"sheet_name": sheet,"round": sel_round, "Initials": Initials, "Approval_Date": Approval_Date, "Percent Female": Female, "Percent Male": Male, "Married Percentage": Married, "Percent w/ Children at Home": Children, "Median Age": Med_Age, "Median HouseHold Income": Med_HHI, "Current_S&O_Notes":omit_notes},ignore_index=True)
        
        return sheet_name_df

    
    def handle_existing_checklists(aud_id,checklists,sel_req_col):
        #temp_checklists=checklists.file_name[checklists['aud_id']==aud_id]
        relevant_checklist=checklists.file_name[checklists['aud_id']==aud_id].loc[max(checklists.file_name[checklists['aud_id']==aud_id].index)]
        #relevant_checklist=temp_checklists.loc[max(temp_checklists.index)]
        existing_tracker = pd.ExcelFile(r'L:\ProductManagement\Standard Audiences\Audience Tracking\Audience Attributes\Omit and Select Checklists\\'+relevant_checklist)
        sheet_names=existing_tracker.sheet_names
        for sheet in sheet_names:
            if "attribute" in sheet.lower():
                relevant_sheet_name=sheet
                sheet_data=pd.read_excel(r'L:\ProductManagement\Standard Audiences\Audience Tracking\Audience Attributes\Omit and Select Checklists\\'+relevant_checklist,sheet_name=relevant_sheet_name,header=10)
                
                sheet_data=sheet_data[~pd.isna(sheet_data['Round'])]
                
                ### newest tracking sheets are recording the selection request ID, so need to check for that column being present
                for col in sheet_data.columns:
                    if "selection req" in col.lower():
                        sel_req_col=True
                sheet_data['Audience Display ID']=aud_id
                sheet_data.rename(columns={"Round" : "round","Current S&O's" : "Current_S&O_Notes", "Approval Date" : "Approval_Date", "% Female" : "Percent Female", "% Male" : "Percent Male", "% Married" : "Married Percentage", "% w/ Children at Home" : "Percent w/ Children at Home", "Median Age" : "Median Age", "Median HHI" : "Median HouseHold Income"}, inplace=True)
                sheet_data=sheet_data[['Audience Display ID', 'round', 'Initials','Current_S&O_Notes', 'Approval_Date', 'Percent Female', 'Percent Male', 'Married Percentage', 'Percent w/ Children at Home', 'Median Age', 'Median HouseHold Income']]
        return sheet_data, sel_req_col
    
    checklists=pd.read_excel(r"C:\Users\gsmittkamp\checklist_files.xlsx",sheet_name='Sheet1')
    aud_attribute_df = pd.DataFrame({"Audience Display ID": [],"sheet_name": [],"round": [], "Initials":[], "Approval_Date": [],"Percent Female": [],"Percent Male": [],"Married Percentage": [],"Percent w/ Children at Home": [],"Median Age": [],"Median HouseHold Income": []})
    
    #create a list of checklists that record selection request_id so that only those are affected by that tracking process
    req_cols = pd.DataFrame({"aud_id": [],"has_req_col": []})
    
    for aud_id in merged.Audience_Display_ID:
        print(aud_id)
        has_sheet=False
        if aud_id not in list(checklists['aud_id'][checklists.has_sheet>0]):
            missing_value=input('''filename for %s is not in list of checklists files saved locally as "checklist_files.xlsx"'''%(aud_id))
        if aud_id in list(checklists['aud_id'][checklists.has_sheet>0]):
            has_sheet=True
            sel_req_col=False
            sheet_data,sel_req_col=handle_existing_checklists(aud_id,checklists,sel_req_col)
            req_cols=req_cols.append({"aud_id":aud_id,"has_req_col": sel_req_col},ignore_index=True)
            ## side task to return the next row number that new attributes will be added to
            #newest_row=sheet_data['round'].max()
            #print("already has tracking sheet for "+aud_id)
            #if aud has attribute tracking sheet, but it hasn't for sure been used for all rounds of selection, ask user to verify
            if len(sheet_data) < 3:
                print('''Attribute Tracking for %s has <3 rounds of selection stored'''%(aud_id))
        #unknown bug causes filenames for UOA's to include "~$" in start of file name, so have to remove those
        check_filenames=existing_files_df[existing_files_df.File_Name.str.lower().str.contains(aud_id.lower())]
        if len(check_filenames) > 1 and "~$" in str(check_filenames.File_Name):
            for file_name in check_filenames.File_Name:
                if "~$" not in str(file_name):
                    correct_file_name=file_name
                    iterable_checklist=existing_files_df.File_Name[existing_files_df.File_Name==correct_file_name]
        else:
            iterable_checklist=existing_files_df.File_Name[existing_files_df.File_Name.str.lower().str.contains(aud_id.lower())]
        if len(iterable_checklist) > 0:
            current_checklist=iterable_checklist.loc[max(iterable_checklist.index)]
            #if there is already a tracking sheet in the workbook, use that for previous stats instead of relying on individual sheets for each round of selection
            if has_sheet==False:
                stop=input("something wrong with attribute file for "+current_checklist+" Ready to proceed? (y/n) \n\nOption: :")
                if stop == 'y':
                
                ################################
                ################################ next code reiterates though above code after allowing user to fix checklist
                
                    print(aud_id)        
                    has_sheet=False
                    if aud_id in list(checklists['aud_id'][checklists.has_sheet>0]):
                        has_sheet=True
                        sheet_data,sel_req_col=handle_existing_checklists(aud_id,checklists)
                        ## side task to return the next row number that new attributes will be added to
                        #newest_row=sheet_data['round'].max()
                        #print("already has tracking sheet for "+aud_id)
                        #if aud has attribute tracking sheet, but it hasn't for sure been used for all rounds of selection, ask user to verify
                        if len(sheet_data) < 3:
                            input('''Attribute Tracking for %s has <3 rounds of selection stored.... Proceed? (y/n) \n\nOption:'''%(aud_id))
                    #unknown bug causes filenames for UOA's to include "~$" in start of file name, so have to remove those
                    check_filenames=existing_files_df[existing_files_df.File_Name.str.lower().str.contains(aud_id.lower())]
                    if len(check_filenames) > 1 and "~$" in str(check_filenames.File_Name):
                        for file_name in check_filenames.File_Name:
                            if "~$" not in str(file_name):
                                correct_file_name=file_name
                                iterable_checklist=existing_files_df.File_Name[existing_files_df.File_Name==correct_file_name]
                    else:
                        iterable_checklist=existing_files_df.File_Name[existing_files_df.File_Name.str.lower().str.contains(aud_id.lower())]
                    if len(iterable_checklist) > 0:
                        current_checklist=iterable_checklist.loc[max(iterable_checklist.index)]
                        
                #########################
                ################################
                
                
                sheet_name_df=get_all_sheets(current_checklist,aud_id)
            else:
                sheet_name_df=sheet_data
            aud_attribute_df=sheet_name_df.append(aud_attribute_df)
            temp_stats=normal_attributes[normal_attributes['Audience Display ID']==aud_id]
            temp_stats['Percent Female']=temp_stats['Percent Female'].str.rstrip('% Female').astype(int)/100
            temp_stats['Percent Male']=temp_stats['Percent Male'].str.rstrip('% Male').astype(int)/100
            temp_stats['sheet_name']=''   
            temp_stats['Initials']='GS'
            temp_stats['Approval_Date']=datetime.datetime.today().strftime('%m/%d/%Y')
            round_count_df=aud_attribute_df[aud_attribute_df['Audience Display ID']==aud_id]
            temp_round_count=int(round_count_df['round'].max())+1
            temp_stats['round']=temp_round_count
            aud_attribute_df=aud_attribute_df.append(temp_stats)
    #aud_attribute_df['Approval_Date']=aud_attribute_df.Approval_Date.str.strip()
    aud_attribute_df['Approval_Date'] = pd.to_datetime(aud_attribute_df['Approval_Date'],errors='coerce')
    aud_attribute_df['Approval_Date'] = pd.to_datetime(aud_attribute_df['Approval_Date'],errors='coerce')
    aud_attribute_df['Current_S&O_Notes'][aud_attribute_df['Current_S&O_Notes']=='']='none'
    aud_attribute_df=aud_attribute_df[['Audience Display ID', 'round', 'Initials','Current_S&O_Notes',
       'Approval_Date', 'Percent Female', 'Percent Male', 'Married Percentage',
       'Percent w/ Children at Home', 'Median Age', 'Median HouseHold Income','sheet_name']]
    aud_attribute_df=aud_attribute_df.sort_values(by=['Audience Display ID','round'],ascending=True)
    aud_attribute_df.rename(columns={'Percent Female': 'Female', 'Percent Male': 'Male', 'Married Percentage':'Married','Percent w/ Children at Home':'Children', 'Median Age':'Age', 'Median HouseHold Income':'HHI'}, inplace=True)
    aud_attribute_df.set_index('Audience Display ID',inplace=True)
    return aud_attribute_df, req_cols


def format_SOs(merged):
    so_formatted=merged
    so_formatted['Initials']='GS'
    so_formatted['Date_Assigned']=datetime.datetime.today()
    so_formatted['SO_Type']='Post'
    so_formatted['Status']='Ready for Approval'
    so_formatted=so_formatted[['Audience_Display_ID','Market_Niche_ID','Initials','Audience_Name','Date_Assigned','SO_Type','Status','Current_S&O_Notes']]
    return so_formatted.set_index('Audience_Display_ID')


def write_to_excel(merged,audience_size,header_data,current_attributes,sos,attribute_evaluation,stat_file_path,option):
    '''Begin Writing each DataFrame to Excel Workbook Sheet'''
    writer = pd.ExcelWriter(stat_file_path, engine='xlsxwriter')
    merged.to_excel(writer, sheet_name='merged_info', index=False)
    audience_size.to_excel(writer, sheet_name='market_size', index=False)
    header_data.to_excel(writer, sheet_name='header_data')
    current_attributes.to_excel(writer, sheet_name='attributes')#, index=False)
    sos.to_excel(writer, sheet_name='sos_formatted')
    if option in ['mb','sb','sa','pr']:
        attribute_evaluation.to_excel(writer, sheet_name='attribute_evaluation')
    writer.save()
    xl = win32.gencache.EnsureDispatch('Excel.Application')
    # wb = xl.Workbooks.Open(stat_file_path)
    # work_sheets=['merged_info','market_size','attributes','sos_formatted','attribute_evaluation']
    # for sheet in work_sheets:
    #     try:
    #         ws = wb.Worksheets(sheet)
    #         ws.Columns.AutoFit()
    #     except:
    #         if option == 'rs':
    #             if sheet != 'attribute_evaluation':
    #                 ws = wb.Worksheets(sheet)
    #                 ws.Columns.AutoFit()
    # Set column widths for one attributes tracking sheet column
    # if option in ['mb','sb','sa','pr']:
    #     ws = wb.Worksheets("attribute_evaluation")
    #     ws.Columns(4).ColumnWidth = 20
    #     wb.Save()
    print("done writing to Excel")
    xl.Workbooks.Open(stat_file_path)
    xl.Visible = True


##### get checklists funtion now only needed with new audiences. Thus not currently being called in main function or initiating processes function
def get_checklist_folder_contents(merged):
    onlyfiles = [f for f in os.listdir(r'L:\ProductManagement\Standard Audiences\Audience Tracking\Audience Attributes\Omit and Select Checklists') if os.path.isfile(os.path.join(r'L:\ProductManagement\Standard Audiences\Audience Tracking\Audience Attributes\Omit and Select Checklists', f))]
    existing_files_df = pd.DataFrame({"File_Name": []})
    for file_name in onlyfiles:
        if "~$" in file_name:
            pass
        else:
            existing_files_df=existing_files_df.append({"File_Name": file_name},ignore_index=True)
    
    #sort existing files dataframe to be ordered same as reverse_sorted
    existing_files_df=existing_files_df.sort_values(by=['File_Name'],ascending=False)
    return reverse_sorted,existing_files_df


def open_excel_files(reverse_sorted,existing_files_df,stat_file_path,tracker_option,open_checklist_option):
    if open_checklist_option == 'yes' or open_checklist_option == 'y':
        new_sheet=input('''Add tracking tabs to checklists? (y/n) \n\nOption: ''')
        xl = win32.gencache.EnsureDispatch('Excel.Application')
        #sort by aud_id so checklists are opened in order
        checklist_count=1
        
        for aud_id in reverse_sorted.Audience_Display_ID:
            print("opening "+aud_id+''' %s of %s'''%(str(checklist_count),str(len(reverse_sorted))))
            iterable_checklist=existing_files_df.File_Name[existing_files_df.File_Name.str.lower().str.contains(aud_id.lower())]
            if len(iterable_checklist) > 0:
                current_checklist=iterable_checklist.loc[max(iterable_checklist.index)]
                template_path = r'"\\Fileshare\Department\ProductManagement\Standard Audiences\Audience Tracking\Audience Attributes\Omit and Select Checklists\TEMPLATE_AUDIENCE ID_MN ID_Demographic Attributes Checklists.xlsx"'
                checklist_path = r'L:\ProductManagement\Standard Audiences\Audience Tracking\Audience Attributes\Omit and Select Checklists\\'+current_checklist
                checklist_wb = xl.Workbooks.Open(checklist_path)
                if new_sheet.lower()=='y':
                    template_workbook = xl.Workbooks.Open(Filename=template_path)
                    template = template_workbook.Worksheets(1)
                    template.Copy(Before=checklist_wb.Worksheets(1))
                #xl.Visible = True
                checklist_count+=1
            else:
                print("currently no checklist file for "+aud_id)
            
               
        xl.Workbooks.Open(stat_file_path)
        #if user chose to open production tracker in main(), open the workbook 
        if tracker_option == 'y' or tracker_option == 'yes':
            print("\n\nopening production tracker")
            xl.Workbooks.Open(r"L:\ProductManagement\Standard Audiences\Audience Tracking\Production Tracking.xlsx")
        xl.Visible = True


def initiate_processing_sequence(shared_wb,option):
    grant_projects=read_audiences(shared_wb,option)
    merged=create_audience_info(grant_projects)
    #merged=select_omits(merged)
    audience_size=check_market_size(merged)
    header_data=create_header_data(merged)
    req_cols=None
    if option in ['mb','sr','sb','sa','pr']:
        current_attributes,normal_attributes=read_Wiland_App_attributess(merged)
    else:
        current_attributes,normal_attributes= pd.DataFrame({"na": []}),pd.DataFrame({"na": []})
    sos=format_SOs(merged)
    print("\nReading Checklist Data... \n\n")
    reverse_sorted=merged.sort_values(by=['Audience_Display_ID'],ascending=False)
    
    from get_filenames import get_file_names
    existing_files_df=get_file_names("L:\ProductManagement\Standard Audiences\Audience Tracking\Audience Attributes\Omit and Select Checklists")
    if option in ['sb','sa','pr']:
        print("\nCreating attribute tracking log...... \n\n")
        attribute_evaluation,req_cols=find_current_attributes(merged,normal_attributes,existing_files_df)
    else:
        attribute_evaluation= pd.DataFrame({"na": []}),pd.DataFrame({"na": []})
    merged.rename(columns={'S&O Notes':'so_notes'}, inplace=True)
    return merged,audience_size,header_data,current_attributes,sos,attribute_evaluation,reverse_sorted,existing_files_df,req_cols

def find_sheet_names(attribute_evaluation,existing_files_df,reverse_sorted):
    aud_data=attribute_evaluation.reset_index()
    aud_data.columns = aud_data.columns.str.lower().str.replace(' ', '_')
    aud_data.rename(columns={'audience_display_id':'aud_id'},inplace=True)
    
    file_names=pd.DataFrame({"Aud_ID": [],"File_Name": [],"Tracking_Sheet_Name": []})
    
    for aud in reverse_sorted.Audience_Display_ID:
        iterable_checklist=existing_files_df.File_Name[existing_files_df.File_Name.str.lower().str.contains(aud.lower())]
        try:
            current_checklist=iterable_checklist.loc[max(iterable_checklist.index)]
        except:
            input("looks like our list does not yet have a record for %s, add it to file named checklist_folder_contents"%aud)
        checklist_path = r'L:\ProductManagement\Standard Audiences\Audience Tracking\Audience Attributes\Omit and Select Checklists\\'+current_checklist
        xli = pd.ExcelFile(checklist_path)
        sheet_names=list(xli.sheet_names)  # see all sheet names
        for sheet in sheet_names:
            if 'att' in sheet.lower() or 'track' in sheet.lower() or 'eval' in sheet.lower():
                right_sheet=sheet
        file_names=file_names.append({"Aud_ID": aud,"File_Name": checklist_path,"Tracking_Sheet_Name": right_sheet},ignore_index=True)
    
    return file_names

       
def find_current_row(attribute_evaluation,reverse_sorted):
    aud_data=attribute_evaluation.reset_index()
    aud_data.columns = aud_data.columns.str.lower().str.replace(' ', '_')
    aud_data.rename(columns={'audience_display_id':'aud_id'},inplace=True)
    
    current_rows=pd.DataFrame({"Aud_ID": [],"Newest_Row": []})
    
    for aud in reverse_sorted.Audience_Display_ID:
        current=aud_data[aud_data.aud_id.str.lower().str.contains(aud.lower())]
        current.reset_index(inplace=True)
        newest_row=current.index.max()+12
        current_rows=current_rows.append({"Aud_ID": aud,"Newest_Row": newest_row},ignore_index=True)
    
    return current_rows


def find_attribute_values(iterable_aud,attribute_evaluation,reverse_sorted):
    aud_data=attribute_evaluation.reset_index()
    aud_data.columns = aud_data.columns.str.lower().str.replace(' ', '_')
    aud_data.rename(columns={'audience_display_id':'aud_id'},inplace=True)
    aud_data=aud_data[aud_data.aud_id.str.lower().str.contains(iterable_aud.lower())]
    
    current=aud_data[aud_data.aud_id.str.lower().str.contains(iterable_aud.lower())]
    current=current[current.index==max(current.index)]
    new_round=max(current['round'])
    initials=current.initials.loc[max(current.index)]
    current_so_notes=current['current_s&o_notes'].loc[max(current.index)]
    approval_date=current.approval_date.loc[max(current.index)]
    female=current.female.loc[max(current.index)]
    male=current.male.loc[max(current.index)]
    married=current.married.loc[max(current.index)]
    children=current.children.loc[max(current.index)]
    age=current.age.loc[max(current.index)]
    hhi=current.hhi.loc[max(current.index)]
        
    return new_round, initials, current_so_notes, approval_date, female, male, married, children, age, hhi


def combine_file_data(attribute_evaluation,reverse_sorted,existing_files_df):
    file_names=find_sheet_names(attribute_evaluation,existing_files_df,reverse_sorted)
    current_rows=find_current_row(attribute_evaluation,reverse_sorted)
    merged_info=pd.merge(file_names,current_rows,how='left',on='Aud_ID')

    return  merged_info

def initiate_atts(attribute_evaluation,reverse_sorted,existing_files_df,req_cols,sleep_option):
    
    
    merged_info=combine_file_data(attribute_evaluation,reverse_sorted,existing_files_df)
    auds_with_req_cols=list(req_cols.aud_id[req_cols.has_req_col>0])
    
    ### TODO ADD TDQM
    for aud in reverse_sorted.Audience_Display_ID:
        #seperate two options here, based on if we are currently tracaking the selection request ids in the attribute evaluation tracker
        if aud not in auds_with_req_cols:
            
            print("editing attribute tracker for "+aud)
            file_name=merged_info.File_Name[merged_info.Aud_ID.str.lower().str.contains(aud.lower())].iloc[0]
            sheet_name=merged_info.Tracking_Sheet_Name[merged_info.Aud_ID.str.lower().str.contains(aud.lower())].iloc[0]
            row=merged_info.Newest_Row[merged_info.Aud_ID.str.lower().str.contains(aud.lower())].iloc[0]
            row=str(int(row))
            iterable_aud=aud
            new_round,initials,current_so_notes,approval_date,female,male,married,children, age, hhi = find_attribute_values(iterable_aud,attribute_evaluation,reverse_sorted)
            # Start by opening the spreadsheet and selecting the main sheet
            workbook = load_workbook(filename=file_name)
            sheet=workbook.get_sheet_by_name(sheet_name)
        
            sheet["B"+row] = new_round
            sheet["C"+row] = initials
            sheet["D"+row] = current_so_notes
            sheet["E"+row] = approval_date
            sheet["F"+row] = female
            sheet["H"+row] = male
            sheet["J"+row] = married
            sheet["L"+row] = children
            sheet["N"+row] = age
            sheet["P"+row] = hhi
            
        if aud in auds_with_req_cols:
            
            #find current selection request id in produciton tracker for that audience
            temp=reverse_sorted[reverse_sorted.Audience_Display_ID.str.lower().str.contains(aud.lower())]
            req_id=temp.Selection_Request_ID.loc[max(temp.index)]
            req_id=int(req_id)
            print("editing attribute tracker for "+aud+" with selection request ID")
            file_name=merged_info.File_Name[merged_info.Aud_ID.str.lower().str.contains(aud.lower())].iloc[0]
            sheet_name=merged_info.Tracking_Sheet_Name[merged_info.Aud_ID.str.lower().str.contains(aud.lower())].iloc[0]
            row=merged_info.Newest_Row[merged_info.Aud_ID.str.lower().str.contains(aud.lower())].iloc[0]
            row=str(int(row))
            iterable_aud=aud
            new_round,initials,current_so_notes,approval_date,female,male,married,children, age, hhi = find_attribute_values(iterable_aud,attribute_evaluation,reverse_sorted)
            # Start by opening the spreadsheet and selecting the main sheet
            workbook = load_workbook(filename=file_name)
            sheet=workbook.get_sheet_by_name(sheet_name)
        
            sheet["B"+row] = new_round
            sheet["C"+row] = req_id
            sheet["D"+row] = initials
            sheet["E"+row] = current_so_notes
            sheet["F"+row] = approval_date
            sheet["G"+row] = female
            sheet["I"+row] = male
            sheet["K"+row] = married
            sheet["M"+row] = children
            sheet["O"+row] = age
            sheet["Q"+row] = hhi
        
        workbook.save(filename=file_name)
        
        if sleep_option.lower()=='y':
            time.sleep(30)
      
        
def aggregated_stats():
    print('\n\n____________CURRENT PRODUCTION STATS____________\n')
    
    shared_wb=pd.read_excel(r'L:\ProductManagement\Standard Audiences\Audience Tracking\Production Tracking.xlsx', header=1,sheet_name='Tracking')
    
    shared_wb=shared_wb [((shared_wb['Initials 12']=='GS')&(((shared_wb.Type=='Refresh')|(shared_wb.Type=='Post'))))|(((shared_wb['Initials 12']=='GS')&(pd.isna(shared_wb['Initials 11'])))&(((shared_wb.Type=='Refresh')|(shared_wb.Type=='Post'))))]
    shared_wb=shared_wb[(shared_wb.Status!='Notified')&(shared_wb.Status!='Shipped')]
    agg_stats=pd.DataFrame(shared_wb.groupby('Status').apply(len)).reset_index()
    status_options={'Model Built': "'mb'", 'Selection Ready': "'sr'", 'Selection Begun': "'sb'", 'Selection Assigned': "'sa'", 'Post-Selection S&O Review': "'pr'", 'Post-Selection S&O Approval':"'pa'", 'Ready to Ship': "'rs'",'Ready for Approval': "-"}
    agg_stats=agg_stats[agg_stats.Status!='Shipped']
    agg_stats['Option']=agg_stats.Status
    agg_stats=agg_stats.replace({"Option": status_options})
    agg_stats=agg_stats.append({"Status": 'Add Selections',"Option": "'new'"},ignore_index=True)
    agg_stats=agg_stats.append({"Status": 'QC Selections',"Option": "'qc'"},ignore_index=True)
    agg_stats.rename(columns={0:'Count'}, inplace=True)
    agg_stats=agg_stats[['Status','Option','Count']]
    
    def pprint_df(dframe):
        print(tabulate(dframe, headers=['Status','Option','Count'], tablefmt='psql', showindex=False))
        
    pprint_df(agg_stats)
    print('\n\n\n')
    


def main():
    #give user option to open production tracker if it is not already open. This variable will be used in same section as checklist file opening below to save lines of extra code
    tracker_option = input('''Open production tracker? (y/n) \n\nOption: ''')
    if tracker_option.lower() not in ('y','n'):
        tracker_option = input('''Open production tracker? (y/n) \n\nOption: ''')
    shared_wb=pd.read_excel(r'L:\ProductManagement\Standard Audiences\Audience Tracking\Production Tracking.xlsx', header=1,sheet_name='Tracking')
    #bring in S&O document to append agg_stats with any audiences ready to ship
        #read in S&O's, note that we drop any rows where at least 8 of the cells are nan
    SOs=pd.read_excel(r'L:\ProductManagement\Standard Audiences\Audience Tracking\Audience Production S&Os.xlsm', header=1,sheet_name='S&O Approval').dropna(axis=0, how='all',thresh=8)
    SOs.rename(columns={"Audience ID" : "Audience Display ID", "MN ID":"Market Niche ID", "Platform Audience Name/MN Description" :"Audience Name","S&O Description":"S&O Notes","S&O Type":"Type"}, inplace=True)
    SOs=SOs[['Audience Display ID', 'Market Niche ID','Initials','Audience Name', 'Type', 'Status','S&O Notes', 'Notes']].fillna(value='')
    #SOs=SOs[SOs.Status!='Ready for Approval']
    SOs.Status[SOs.Status=="Approved - No Additional S&Os"]="Time to Ship"
    SOs.Status[SOs.Status=="Approved - Additional S&Os"]="Apply New S&O's"
    shared_wb=shared_wb.append(SOs)
    #narrow down audiences to those that are Grant's, that are in relevant status (retired vs refresh), or are pulled in from the SO's document Karen interacts with
    #shared_wb=shared_wb[(shared_wb['Initials 11']=='GS')&(((shared_wb.Type=='Refresh')|(shared_wb.Type=='Post')))|(((shared_wb.Status=='Ready to Ship')&(shared_wb['Initials 11']=='GS')))|(shared_wb['Initials 11']=='GS')&(((shared_wb['Initials 12']=='GS')]
    shared_wb=shared_wb[(shared_wb['Initials 12']=='GS')&(((shared_wb.Type=='Refresh')|(shared_wb.Type=='Post')))|(((shared_wb.Status=='Ready to Ship')&(shared_wb['Initials 12']=='GS')))]
    shared_wb=shared_wb[(shared_wb.Status!='Notified')&(shared_wb.Status!='Shipped')]
    #remove any non USP or UAA audience from consideration as they have seperate workflow (UOA's have no attribute tracking)
    if "uoa" in str(shared_wb['Audience Display ID']).lower():
        special_option = input('''UOA's found in aggregated stats, proceed WITHOUT UOA's? (y/n) \n\nOption: ''')
        if special_option.lower() != 'y':
            shared_wb=shared_wb[shared_wb['Audience Display ID'].str.lower().str.contains('uoa')]
    aggregated_stats()
    option = input('''Choose an option from the chart above with corresponding status, or type 'quit' to cancel\n\nOption: ''')
    if option not in ('mb','sr','sb','sa','pr','pa','rs','quit','new'):
        option = input('''Input unrecognized - Choose an option from the chart above with corresponding status, or type 'quit' to cancel\n\nOption: ''')
    option=option.lower()
    if option=='new':
        new_selections.main()
        qty=len(shared_wb[shared_wb.Status=='Selection Ready'])
        pass
    elif option in ['mb','sr','sb','sa','pr','pa','rs','new']:
        merged,audience_size,header_data,current_attributes,sos,attribute_evaluation,reverse_sorted,existing_files_df,req_cols=initiate_processing_sequence(shared_wb,option)
        attribute_option = input('''Edit attribute tracking sheets and save them?\n\nOption: ''')
        #beacuse of potential negative impact, ensure use does want to edit the tracking sheets
        if attribute_option == 'y' or attribute_option == 'yes':
            attribute_option_2 = input('''ARE YOU SURE??? Edit and save them?\n\nOption: ''')
            if attribute_option_2 == 'y' or attribute_option == 'yes':
                sleep_option = input('''Pause in between checklists? (y/n) \n\nOption: ''')
                initiate_atts(attribute_evaluation,reverse_sorted,existing_files_df,req_cols,sleep_option)
        #add_worksheet(existing_files_df,merged)
        stat_file_name=datetime.datetime.today().strftime('%Y.%m.%d')+'_'+datetime.datetime.now().strftime("%H.%M")+'_Aud_Prod_Stats.xlsx' 
        stat_file_path=r'C:\Users\gsmittkamp\Audience_Production_Stats\%s'%(stat_file_name)
        write_to_excel(merged,audience_size,header_data,current_attributes,sos,attribute_evaluation,stat_file_path,option)
        if option in ['sb','sa','pr','pa','rs']:
            open_checklist_option = input('''Open %s checklist files? (y/n) \n\nOption: '''%(str(len(merged))))
        else:
            open_checklist_option = 'n'
        open_excel_files(reverse_sorted,existing_files_df,stat_file_path,tracker_option,open_checklist_option)
        qty=len(merged)

    elif option=='quit':
        print("cancelling process")
        pass
    else:
        print("input unrecognized")
        pass

    print("complete, have a great rest of the day")
    return option, qty


if __name__ == '__main__':
    clear_excel_cache()
    option=main()
    #start by gathering aggregated production stats to give user insight intro statusoptions before doing further processing
#    try:
#        main()
#    except:
#        excel_chache = input('''Encounted Error, clear Excel cache and re-attempt? (y/n) \n\nOption: ''')
#        if excel_chache == 'y':
#            clear_excel_cache()
#            main()
                

#def create_backup():
#    shared_wb=pd.read_excel(r'L:\ProductManagement\Standard Audiences\Audience Tracking\Production Tracking.xlsx', header=1,sheet_name='Tracking')
#    writer = pd.ExcelWriter(r'C:\Users\gsmittkamp\production_backup.xlsx', engine='xlsxwriter')
#    shared_wb.to_excel(writer, sheet_name='backup', index=False)
#    writer.save()
#    return shared_wb

#todo find most recent projects to work on
#shared_wb=pd.read_excel(r'BACKUP Production Tracking.xlsx',header=0, sheet_name='Tracking')
#min_recent_date = shared_wb['Latest Start Date'].min()
#newest_auds=shared_wb[shared_wb['Latest Start Date']==min_recent_date]
#next_up=shared_wb[newest_auds.index.max()+1:newest_auds.index.max()+11]
#df['StartDate'] = pd.to_datetime(df['StartDate'],errors='coerce')
#>> least_recent_date = df['StartDate'].min()
#>> recent_date = df['StartDate'].max()
